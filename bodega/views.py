from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.db.models import F
from .models import *
from usuarios.models import PerfilUsuario
from .forms import SalidaProductoForm, StockForm, EntregaCorteForm, InsumoForm, IngresoInsumoForm, UsoInsumoForm
import json
import openpyxl
import os
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook

# Vista del inventario de bodega
@login_required(login_url='login')
def inventario_bodega(request):
    stock_queryset = Stock.objects.select_related('producto_variante__producto', 'producto_variante__color', 'producto_variante__talla', 'producto_variante__diseno')

    stock_list = []
    for item in stock_queryset:
        entradas = EntregaCorte.objects.filter(producto=item.producto_variante).exists()
        salidas = SalidaProducto.objects.filter(producto__producto_variante=item.producto_variante).exists()
        tiene_movimientos = entradas or salidas

        stock_list.append({
            'stock': item,
            'tiene_movimientos': tiene_movimientos,
        })

    return render(request, 'bodega/inventario-bodega.html', {'stock': stock_list})

def crear_stock(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            producto_variante = form.cleaned_data['producto_variante']
            cantidad = form.cleaned_data['cantidad']

            # Validar si ya existe
            if Stock.objects.filter(producto_variante=producto_variante).exists():
                messages.error(request, "Ya existe este producto en bodega.")
            else:
                form.save()
                messages.success(request, "Producto agregado al inventario de bodega.")
                return redirect('inventario_bodega')
    else:
        form = StockForm()

    return render(request, 'bodega/form_stock.html', {
        'form': form,
        'accion': 'Agregar al inventario'
    })

def editar_stock(request, stock_id):
    stock = get_object_or_404(Stock, id=stock_id)

    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock)
        if form.is_valid():
            form.save()
            messages.success(request, "Inventario actualizado.")
            return redirect('inventario_bodega')
    else:
        form = StockForm(instance=stock)

    return render(request, 'bodega/form_stock.html', {
        'form': form,
        'accion': 'Editar inventario'
    })

@login_required(login_url='login')
@require_POST
def eliminar_stock(request, stock_id):
    if request.method == "POST":
        stock = get_object_or_404(Stock, id=stock_id)
        stock.delete()
        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Método no permitido"})

# Vista de entradas de productos
@login_required(login_url='login')
def entradas_producto(request):
    entrega_corte = EntregaCorte.objects.all()
    return render(request, 'bodega/entradas-productos.html', {'entrega_corte': entrega_corte})

@login_required
def crear_entrada(request):
    if request.method == 'POST':
        form = EntregaCorteForm(request.POST)
        if form.is_valid():
            entrada = form.save(commit=False)
            try:
                entrada.user_responsable = request.user.perfilusuario
            except PerfilUsuario.DoesNotExist:
                messages.error(request, "Este usuario no tiene un perfil asociado.")
                return redirect('entradas')
            entrada.save()
            messages.success(request, "Entrada registrada exitosamente.")
            return redirect('entradas')
    else:
        form = EntregaCorteForm()
    return render(request, 'bodega/form_entrada.html', {
        'form': form,
        'accion': 'Registrar'
    })

@login_required
def editar_entrada(request, entrada_id):
    entrada = get_object_or_404(EntregaCorte, id=entrada_id)
    if request.method == 'POST':
        form = EntregaCorteForm(request.POST, instance=entrada)
        if form.is_valid():
            form.save()
            messages.success(request, "Entrada actualizada correctamente.")
            return redirect('entradas')
    else:
        form = EntregaCorteForm(instance=entrada)
    return render(request, 'bodega/form_entrada.html', {
        'form': form,
        'accion': 'Editar'
    })

@login_required
def eliminar_entrada(request, entrada_id):
    entrada = get_object_or_404(EntregaCorte, id=entrada_id)
    devolver = request.GET.get('devolver') == '1'

    with transaction.atomic():
        cantidad_en_bodega = entrada.cantidad - entrada.cantidad_lavado

        try:
            stock = Stock.objects.select_for_update().get(producto_variante=entrada.producto)

            if devolver:
                stock.cantidad = F('cantidad') - cantidad_en_bodega
                stock.save()
        except Stock.DoesNotExist:
            pass

        if entrada.cantidad_lavado > 0:
            salida_lavado = SalidaProducto.objects.filter(
                producto__producto_variante=entrada.producto,
                cantidad=entrada.cantidad_lavado,
                es_lavado=True,
                user_responsable=entrada.user_responsable
            ).first()
            if salida_lavado:
                salida_lavado.delete()

        entrada.delete()

    return JsonResponse({'success': True})

# Vista de salidas de productos
@login_required(login_url='login')
def salidas_producto(request):
    salida_producto = SalidaProducto.objects.all() 
    return render(request, 'bodega/salidas-productos.html', {'salida_producto': salida_producto})

@login_required(login_url='login')
def crear_salida_producto(request):
    if request.method == 'POST':
        form = SalidaProductoForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salida registrada correctamente.')
            return redirect('salidas_producto')
    else:
        form = SalidaProductoForm(request=request)

    return render(request, 'bodega/form_salida.html', {
        'form': form,
        'accion': 'Registrar salida'
    })

@login_required(login_url='login')
def editar_salida_producto(request, salida_id):
    salida = get_object_or_404(SalidaProducto, id=salida_id)

    if request.method == 'POST':
        form = SalidaProductoForm(request.POST, instance=salida, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salida actualizada correctamente.')
            return redirect('salidas_producto')
    else:
        form = SalidaProductoForm(instance=salida, request=request)

    return render(request, 'bodega/form_salida.html', {
        'form': form,
        'accion': 'Editar salida'
    })

@csrf_exempt
@login_required(login_url='login')
@require_POST
def eliminar_salida_producto(request, salida_id):
    try:
        salida = get_object_or_404(SalidaProducto, id=salida_id)

        data = json.loads(request.body)
        devolver = data.get("devolver", False)

        if devolver and not salida.es_lavado:
            # Devolver al stock
            stock = salida.producto
            stock.cantidad += salida.cantidad
            stock.save()

        salida.delete()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
def historial_movimientos_producto(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante)
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante)

    context = {
        'variante': variante,
        'entradas': entradas,
        'salidas': salidas,
    }
    return render(request, 'bodega/historial_movimientos_producto.html', context)

@login_required
def exportar_historial_excel(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante).order_by('-fecha')
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante).order_by('-fecha')

    if not entradas.exists() and not salidas.exists():
        return render(request, 'bodega/sin_movimientos.html', status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Información del producto
    ws.append(["Producto:", str(variante.producto.referencia)])
    ws.append(["Color:", str(variante.color)])
    ws.append(["Talla:", str(variante.talla)])
    ws.append(["Diseño:", str(variante.diseno.nombre) if variante.diseno else "Sin diseño"])
    ws.append([])  # Línea en blanco

    # Encabezados
    ws.append(["Tipo", "Fecha", "Cantidad", "Responsable", "Estado"])

    for entrada in entradas:
        ws.append([
            "Entrada",
            entrada.fecha.strftime('%Y-%m-%d'),
            entrada.cantidad,
            str(entrada.user_responsable),
            "Completado"
        ])

    for salida in salidas:
        ws.append([
            "Lavado" if salida.es_lavado else "Salida",
            salida.fecha.strftime('%Y-%m-%d'),
            salida.cantidad,
            str(salida.user_responsable),
            salida.estado
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Historial_{variante_id}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def exportar_historial_pdf(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante).order_by('-fecha')
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante).order_by('-fecha')

    if not entradas.exists() and not salidas.exists():
        return render(request, 'bodega/sin_movimientos.html', status=400)

    # Configurar respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Historial_{variante_id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # --- Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'outfitcore', 'static', 'images', 'logo-img.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=2 * inch, height=1 * inch)
        elements.append(logo)

    elements.append(Spacer(1, 12))

    # --- Título ---
    title = Paragraph(f"<strong>Historial de Movimientos</strong>", styles['Title'])
    elements.append(title)

    # --- Info del producto ---
    producto_info = f"""
        <strong>Referencia:</strong> {variante.producto.referencia}<br/>
        <strong>Color:</strong> {variante.color}<br/>
        <strong>Talla:</strong> {variante.talla}<br/><br/>
    """
    elements.append(Paragraph(producto_info, styles['Normal']))

    # --- Tabla de movimientos ---
    data = [['Tipo', 'Fecha', 'Cantidad', 'Responsable', 'Estado']]

    for entrada in entradas:
        data.append([
            'Entrada',
            entrada.fecha.strftime('%Y-%m-%d'),
            entrada.cantidad,
            str(entrada.user_responsable),
            'Completado'
        ])

    for salida in salidas:
        tipo = 'Lavado' if salida.es_lavado else 'Salida'
        data.append([
            tipo,
            salida.fecha.strftime('%Y-%m-%d'),
            salida.cantidad,
            str(salida.user_responsable),
            salida.estado
        ])

    if len(data) == 1:
        elements.append(Paragraph("No hay movimientos registrados para este producto.", styles['Normal']))
    else:
        table = Table(data, colWidths=[1.2*inch]*5)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)

    # --- Generar PDF ---
    doc.build(elements)
    return response

def inventario_insumos(request):
    insumos = Insumo.objects.all().order_by('nombre')
    return render(request, 'bodega/insumos/inventario_insumos.html', {'insumos': insumos}) 

def agregar_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre']
            if Insumo.objects.filter(nombre__iexact=nombre).exists():
                messages.error(request, 'Ya existe un insumo con ese nombre.')
            else:
                form.save()
                messages.success(request, 'Insumo creado exitosamente.')
                return redirect('inventario_insumos')
    else:
        form = InsumoForm()
    
    return render(request, 'bodega/insumos/form_insumo.html', {
        'form': form,
        'accion': 'Agregar insumo'
    })

def editar_insumo(request, insumo_id):
    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Insumo actualizado correctamente.')
            return redirect('inventario_insumos')
    else:
        form = InsumoForm(instance=insumo)
    
    return render(request, 'bodega/insumos/form_insumo.html', {
        'form': form,
        'accion': 'Editar insumo'
    })

@login_required
def eliminar_insumo(request, insumo_id):
    insumo = get_object_or_404(Insumo, id=insumo_id)

    with transaction.atomic():
        insumo.delete()

    return JsonResponse({'success': True})

# Ingreso Insumo

@login_required
def listar_ingresos_insumo(request):
    ingresos = IngresoInsumo.objects.select_related('user_responsable').order_by('-fecha')
    return render(request, 'bodega/insumos/listar_ingresos_insumo.html', {'ingresos': ingresos})

@login_required
@transaction.atomic
def agregar_ingreso_insumo(request):
    if request.method == 'POST':
        form = IngresoInsumoForm(request.POST)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.user_responsable = request.user.perfilusuario
            ingreso.save()  # Guardamos el ingreso

            # Crear automáticamente el detalle del ingreso
            insumo = ingreso.insumo
            cantidad = ingreso.cantidad

            DetalleIngresoInsumo.objects.create(
                ingreso=ingreso,
                insumo=insumo,
                cantidad=cantidad
            )

            # Actualizar el stock del insumo
            insumo.stock_actual += cantidad
            insumo.save()

            return redirect('ingresos_insumos')  # URL a la lista de ingresos
    else:
        form = IngresoInsumoForm()

    context = {
        'form': form,
        'accion': 'Crear',
    }
    return render(request, 'bodega/insumos/form_ingreso_insumo.html', context)

@login_required
def editar_ingreso_insumo(request, ingreso_id):
    ingreso = get_object_or_404(IngresoInsumo, id=ingreso_id)
    cantidad_anterior = ingreso.cantidad  # Guardamos la cantidad original

    if request.method == 'POST':
        form = IngresoInsumoForm(request.POST, instance=ingreso)
        if form.is_valid():
            ingreso_editado = form.save(commit=False)
            ingreso_editado.user_responsable = request.user.perfilusuario  # Asegurar el user_responsable

            # Actualizar el stock restando lo viejo y sumando lo nuevo
            diferencia = ingreso_editado.cantidad - cantidad_anterior
            ingreso_editado.insumo.stock_actual += diferencia
            ingreso_editado.insumo.save()

            ingreso_editado.save()
            messages.success(request, 'Ingreso actualizado correctamente.')
            return redirect('ingresos_insumos')
    else:
        form = IngresoInsumoForm(instance=ingreso)

    return render(request, 'bodega/insumos/form_ingreso_insumo.html', {
        'form': form,
        'accion': 'Editar ingreso de insumo'
    })


@login_required
def eliminar_ingreso_insumo(request, ingreso_id):
    ingreso = get_object_or_404(IngresoInsumo, id=ingreso_id)

    with transaction.atomic():
        # Restar el stock
        for detalle in ingreso.detalles.all():
            insumo = detalle.insumo
            insumo.stock_actual -= detalle.cantidad
            insumo.save()

        # Eliminar el ingreso (esto también debería eliminar los detalles si usas on_delete=models.CASCADE)
        ingreso.delete()

    return JsonResponse({'success': True})


def detalle_ingreso_insumo(request, ingreso_id):
    ingreso = get_object_or_404(IngresoInsumo, id=ingreso_id)
    detalles = ingreso.detalles.all()
    return render(request, 'bodega/insumos/detalle_ingreso_insumo.html', {
        'ingreso': ingreso,
        'detalles': detalles
    })

@login_required
def exportar_detalle_ingreso_excel(request, ingreso_id):
    ingreso = get_object_or_404(IngresoInsumo, id=ingreso_id)
    detalles = DetalleIngresoInsumo.objects.filter(ingreso=ingreso)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Ingreso"

    # Información general
    ws.append(["Ingreso de Insumos"])
    ws.append(["Fecha:", ingreso.fecha.strftime('%d/%m/%Y')])
    ws.append(["Proveedor:", ingreso.proveedor.Razon_Social])
    ws.append(["Registrado por:", str(ingreso.user_responsable)])
    ws.append(["Estado:", ingreso.estado])
    ws.append([])

    # Encabezado tabla
    ws.append(["#", "Insumo", "Cantidad", "Unidad de Medida", "Tipo de Insumo"])

    for i, detalle in enumerate(detalles, start=1):
        ws.append([
            i,
            detalle.insumo.nombre,
            detalle.cantidad,
            detalle.insumo.unidad_medida.nombre,
            detalle.insumo.tipo_insumo.nombre,
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Preparar respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=Detalle_Ingreso_{ingreso_id}.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_detalle_ingreso_pdf(request, ingreso_id):
    ingreso = get_object_or_404(IngresoInsumo, id=ingreso_id)
    detalles = DetalleIngresoInsumo.objects.filter(ingreso=ingreso)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Detalle_Ingreso_{ingreso_id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Logo (opcional)
    logo_path = os.path.join(settings.BASE_DIR, 'outfitcore', 'static', 'images', 'logo-img.png')
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=2 * inch, height=1 * inch))

    elements.append(Spacer(1, 12))

    # Título e información general
    elements.append(Paragraph("Detalle de Ingreso de Insumos", styles['Title']))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<strong>Fecha:</strong> {ingreso.fecha.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Paragraph(f"<strong>Proveedor:</strong> {ingreso.proveedor.Razon_Social}", styles['Normal']))
    elements.append(Paragraph(f"<strong>Registrado por:</strong> {ingreso.user_responsable}", styles['Normal']))
    elements.append(Paragraph(f"<strong>Estado:</strong> {ingreso.estado}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tabla de insumos
    data = [["#", "Insumo", "Cantidad", "Unidad de Medida", "Tipo de Insumo"]]
    for i, d in enumerate(detalles, 1):
        data.append([
            i,
            d.insumo.nombre,
            d.cantidad,
            d.insumo.unidad_medida.nombre,
            d.insumo.tipo_insumo.nombre
        ])

    table = Table(data, colWidths=[0.5*inch, 2.2*inch, 1*inch, 1.5*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

@login_required(login_url='login')
def lista_uso_insumos(request):
    usos = UsoInsumo.objects.select_related('insumo', 'producto', 'user_responsable')
    return render(request, 'bodega/insumos/uso_insumo_list.html', {'usos': usos})

@login_required
def obtener_stock_insumo(request, insumo_id):
    try:
        insumo = Insumo.objects.get(id=insumo_id)
        return JsonResponse({'stock_actual': insumo.stock_actual})
    except Insumo.DoesNotExist:
        return JsonResponse({'stock_actual': 0})

@login_required(login_url='login')
def crear_uso_insumo(request):
    if request.method == 'POST':
        form = UsoInsumoForm(request.POST)
        if form.is_valid():
            uso = form.save(commit=False)

            if uso.cantidad > uso.insumo.stock_actual:
                form.add_error('cantidad', 'No hay suficiente stock del insumo.')

            uso.user_responsable = request.user.perfilusuario
            uso.insumo.stock_actual -= uso.cantidad
            uso.insumo.save()
            form.save()
            messages.success(request, 'Agregado correctamente')
            return redirect('lista_uso_insumos')  # Asegúrate que esta URL exista
    else:
        form = UsoInsumoForm()

    context = {
        'form': form,
        'accion': 'Registrar uso de insumo',
    }
    return render(request, 'bodega/insumos/uso_insumo_form.html', context)


@login_required(login_url='login')
@transaction.atomic
def editar_uso_insumo(request, uso_id):
    uso = get_object_or_404(UsoInsumo, id=uso_id)
    insumo = uso.insumo  # Guardamos la referencia del insumo original
    cantidad_anterior = uso.cantidad  # Para calcular la diferencia luego

    if request.method == 'POST':
        form = UsoInsumoForm(request.POST, instance=uso)
        if form.is_valid():
            uso_editado = form.save(commit=False)
            nueva_cantidad = uso_editado.cantidad
            diferencia = nueva_cantidad - cantidad_anterior

            if diferencia > 0 and diferencia > insumo.stock_actual:
                form.add_error('cantidad', 'No hay suficiente stock del insumo para este ajuste.')
            else:
                # Ajustar el stock
                insumo.stock_actual -= diferencia
                insumo.save()
                uso_editado.save()
                messages.success(request, "Uso de insumo actualizado.")
                return redirect('lista_uso_insumos')
    else:
        form = UsoInsumoForm(instance=uso)

    return render(request, 'bodega/insumos/uso_insumo_form.html', {
        'form': form,
        'accion': 'Editar uso de insumo'
    })


@csrf_exempt
@login_required(login_url='login')
@require_POST
def eliminar_uso_insumo(request, uso_id):
    uso = get_object_or_404(UsoInsumo, id=uso_id)
    devolver = request.GET.get('devolver') == '1'

    try:
        with transaction.atomic():
            if devolver:
                # Devolver la cantidad al stock del insumo
                insumo = Insumo.objects.select_for_update().get(pk=uso.insumo.pk)
                insumo.stock_actual = F('stock_actual') + uso.cantidad
                insumo.save()

            # Eliminar el uso
            uso.delete()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



